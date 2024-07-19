from datetime import datetime
import re
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog

import PyPDF2
import openpyxl
from openpyxl.styles import PatternFill
import requests



workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'Номер заказа'
sheet['B1'] = '№ Заказа на работу'
sheet['C1'] = 'Дата'
sheet['D1'] = 'Сумма'
sheet['E1'] = 'Базовая станция'

red_fill = PatternFill(start_color='FFFF0000',
                    end_color='FFFF0000',
                    fill_type='solid')


def browse_file_1():
    file_path = filedialog.askopenfilename()
    entry_1.delete(0, tk.END)
    entry_1.insert(0, file_path)

def process_files():
    file_path = entry_1.get()
    wb_seven = openpyxl.load_workbook(file_path, data_only=True)
    seven_sheet = wb_seven.worksheets[0]
    sheet_column_U = seven_sheet['U']
    pattern = r'"([^"]*)"'


    # Указываем путь к папке, содержащей PDF-файлы
    pdf_path = './pdf'

    # Получаем список всех файлов в указанной папке
    pdfs = os.listdir(pdf_path)
    index = 2


    for pdf_name in pdfs:
        print(pdf_name)
        with open(f'./pdf/{pdf_name}', 'rb') as pdf_file:

            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)

            pattern_order = r'Заказ № (\d+)'
            pattern_date = r'(\d{2}\.\d{2})\.(\d{2})'

            page = pdf_reader.pages[0].extract_text()
            order_num = re.search(pattern_order, page).group(1)
            year = re.search(pattern_date, page).group(2)
            year = "20"+ year

            date = re.search(pattern_date, page).group(1) + '.' + year
            # kt_res = re.findall(r'\[\s*([A-Z0-9/]+\s*)\]', page)
            # #print(kt_res)
            # Kt = [kti for kti in kt_res if re.match(r'[A-Z]{2}/\d{4}/[A-Z\d]+', kti)]
            # print(Kt)
            # kt_order = f'{order_num}/{Kt[0]}'

            BS_list = []
            sum = []
            KT_res = []

            for i in range(num_pages):
                page = pdf_reader.pages[i].extract_text()
                res = page.split('\n')
                for j in res:
                    kt_res = re.findall(r'\[\s*([A-Z0-9\/]+\s*)\]', j)
                    Kt = [kti for kti in kt_res if re.match(r'[A-Z]{2}/\d{4}/[A-Z\d]+', kti)]
                    if Kt:
                        KT_res.append(Kt[0])
                    if "бс \"" in j.lower() or "т бс " in j.lower() or "ты бс" in j.lower() or " БС№" in j:
                        BS_pos = j.find("БС")
                        BS_list.append(j[BS_pos:])
                    elif "бот ррл" in j.lower() or "боты ррл" in j.lower():
                        BS_pos = j.find("РРЛ")
                        BS_list.append(j[BS_pos:])
                    # elif " ТО" in j:
                    #     BS_pos = j.find("ТО")
                    #     BS_list.append(j[BS_pos:])

                    if "всего с учетом ндс:" in j.lower():
                        sum.append(j[20:])

            for i in range(len(sum)):
                kt_order = f'{order_num}/{KT_res[i]}'
                sheet[f'B{index}'] = kt_order
                sheet[f'C{index}'] = date
                sheet[f'D{index}'] = sum[i]
                sheet[f'E{index}'] = BS_list[i]

                BS_for_comment = None
                if '"' in BS_list[i]:
                    match = re.search(pattern, BS_list[i])
                    if match:
                        BS_for_comment = match.group(1)

                for j in range(len(sheet_column_U)):
                    if sheet_column_U[j].value is not None and BS_for_comment and BS_for_comment in sheet_column_U[j].value:
                        sum_ = re.sub(r'\s*', '', sum[i])
                        if seven_sheet[f'Y{j+1}'].value == float(sum_):
                            sheet[f'A{index}'] = seven_sheet[f'A{j+1}'].value
                    elif sheet_column_U[j].value is not None and BS_list[i] in sheet_column_U[j].value:
                        sum_ = re.sub(r'\s*', '', sum[i])
                        if seven_sheet[f'Y{j+1}'].value == float(sum_):
                            sheet[f'A{index}'] = seven_sheet[f'A{j+1}'].value

                if sheet[f'A{index}'].value is None or sheet[f'A{index}'].value == '':
                    for fill_i in range(1, 6):
                        sheet.cell(row=sheet.max_row, column=fill_i).fill = red_fill
                    sheet[f'A{index}'] = "NULL"
                index += 1

    workbook.save('Результат.xlsx')

    workbook.close()
    def send_report(text=None, process=None, responsible=None):

        requests.post(f"https://script.google.com/macros/s/AKfycbzDwjE6Pu1a7otho2EHwbI-4yNoEmLijTfwWfI3toWpDpJ6rc-O1pKljV6XMLJmQIyJ/exec?time={datetime.now().strftime('%d.%m.%Y %H:%M:%S')}&process={process}&responsible={responsible}&text={text}")

    send_report(text="Обработчик_PDF",process="Обработчик_PDF_Сверка_7.15.2", responsible=os.getlogin())
    result_label.config(text=f"ФАЙЛ ГОТОВ!")

print("id 1_4")
root = tk.Tk()
root.title("Выбор файлов")
root.geometry("500x200")

file_frame = ttk.LabelFrame(root, text="Выбор файлов")
file_frame.grid(column=0, row=0, padx=10, pady=10)


label_2 = tk.Label(file_frame, text="Выберите файл 7.15.2:", width=20)
label_2.grid(row=1, column=0, padx=5, pady=5)

entry_1 = tk.Entry(file_frame, width=30)
entry_1.grid(row=1, column=1, padx=5, pady=5)

browse_button_2 = ttk.Button(file_frame, text="Обзор", command=browse_file_1)
browse_button_2.grid(row=1, column=2, padx=5, pady=5)


process_button = ttk.Button(root, text="Обработать файлы", command=process_files, width=20)
process_button.grid(row=2, column=0, columnspan=3, padx=5, pady=10)

# Результат извлечения
result_label = ttk.Label(root, text="")
result_label.grid(column=0, row=5, columnspan=3, padx=10, pady=10)

root.mainloop()